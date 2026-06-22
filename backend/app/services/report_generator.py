import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from typing import List, Dict, Any

class SecurityReportGenerator:
    @staticmethod
    def generate_pdf_report(tenant_name: str, events: List[Dict[str, Any]], ai_summary: str) -> bytes:
        """
        توليد تقرير أمني رسمي بصيغة PDF يتضمن التحليل الأوتوماتيكي بالذكاء الاصطناعي وجدول الأحداث
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=letter,
            rightMargin=40, leftMargin=40, 
            topMargin=40, bottomMargin=40
        )
        
        styles = getSampleStyleSheet()
        
        # إنشاء نمط للغة العربية (RTL)
        arabic_style = ParagraphStyle(
            'ArabicText',
            parent=styles['Normal'],
            fontName='Helvetica', # في الإنتاج يفضل استخدام خط يدعم العربية مثل Amiri أو Scheherazade
            fontSize=10,
            leading=14,
            alignment=2 # Alignment 2 is RIGHT
        )
        
        title_style = ParagraphStyle(
            'ArabicTitle',
            parent=styles['Title'],
            fontName='Helvetica-Bold',
            fontSize=16,
            leading=20,
            alignment=1 # Center
        )

        elements = []
        
        # 1. عنوان التقرير
        elements.append(Paragraph(f"التقرير الأمني والتشغيلي لمنشأة: {tenant_name}", title_style))
        elements.append(Spacer(1, 20))
        
        # 2. ملخص الذكاء الاصطناعي (AI Summary)
        elements.append(Paragraph("<b>أولاً: التحليل والملخص التنفيذي بالذكاء الاصطناعي:</b>", arabic_style))
        elements.append(Spacer(1, 8))
        elements.append(Paragraph(ai_summary, arabic_style))
        elements.append(Spacer(1, 25))
        
        # 3. جدول الأحداث والمخالفات
        elements.append(Paragraph("<b>ثانياً: سجل الأحداث والتنبيهات المرصودة:</b>", arabic_style))
        elements.append(Spacer(1, 10))
        
        # إعداد بيانات الجدول
        # العناوين بالترتيب المعكوس لتناسب العرض العربي من اليمين لليسار
        table_data = [["تاريخ الحدث", "الهوية المتوقعة", "نوع الكائن", "الكاميرا", "رقم الحدث"]]
        
        for idx, event in enumerate(events):
            table_data.append([
                event.get("created_at", "").split("T")[0],
                event.get("culprit_detected", "مجهول"),
                event.get("event_type", "person"),
                event.get("camera_name", "Camera"),
                str(idx + 1)
            ])
            
        # إعداد نمط الجدول
        t = Table(table_data, colWidths=[120, 100, 100, 100, 50])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#2B2E4A")),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('BACKGROUND', (0,1), (-1,-1), colors.HexColor("#F4F4F2")),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,1), (-1,-1), 9),
        ]))
        
        elements.append(t)
        
        # بناء المستند
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def generate_excel_report(tenant_name: str, events: List[Dict[str, Any]], ai_summary: str) -> bytes:
        """
        توليد تقرير أمني واحترافي بصيغة Excel
        """
        wb = Workbook()
        
        # الورقة الأولى: الملخص التنفيذي
        ws1 = wb.active
        ws1.title = "Executive Summary"
        ws1.views.sheetView[0].showGridLines = True
        
        ws1.column_dimensions['A'].width = 15
        ws1.column_dimensions['B'].width = 80
        
        # الهيدر
        ws1['A1'] = "تقرير منشأة"
        ws1['B1'] = tenant_name
        ws1['A1'].font = Font(bold=True, size=12)
        ws1['B1'].font = Font(bold=True, size=12, color="2B2E4A")
        
        # ملخص الذكاء الاصطناعي
        ws1['A3'] = "ملخص الـ AI"
        ws1['B3'] = ai_summary
        ws1['A3'].font = Font(bold=True)
        ws1['B3'].alignment = Alignment(wrap_text=True, vertical="top")
        ws1.row_dimensions[3].height = 150
        
        # الورقة الثانية: جدول الأحداث التفصيلي
        ws2 = wb.create_sheet(title="Events Detail")
        ws2.views.sheetView[0].showGridLines = True
        
        # العناوين
        headers = ["Event ID", "Camera Name", "Event Type", "Culprit / Person", "Time Recorded", "Description"]
        ws2.append(headers)
        
        # تنسيق العناوين
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2B2E4A", end_color="2B2E4A", fill_type="solid")
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws2.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            
        # إضافة البيانات
        for event in events:
            ws2.append([
                event.get("event_id", ""),
                event.get("camera_name", ""),
                event.get("event_type", ""),
                event.get("culprit_detected", "Unknown"),
                event.get("created_at", ""),
                event.get("raw_description", "")
            ])
            
        # ضبط عرض الأعمدة تلقائياً
        for col in ws2.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
